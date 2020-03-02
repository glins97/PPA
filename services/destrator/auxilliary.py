import os
import os.path
import pickle
import codecs 
import io
import re

from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

import openpyxl 
import pandas
import numpy as np
import os 
from copy import copy, deepcopy
import datetime
from matplotlib import pyplot

def get_row(cell):
    return int(re.search(r'(\d+?)$', cell).group(1))
    
def name_format(fn):
    return ' TPS '.join(fn.split('TPS')).replace('(respostas)', '').strip()

def auth():
    SCOPES = ['https://www.googleapis.com/auth/drive']
    creds = None
    if os.path.exists('./destrator/credentials/token'):
        with open('./destrator/credentials/token', 'rb') as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                './destrator/credentials/client_secret.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('./destrator/credentials/token', 'wb') as token:
            pickle.dump(creds, token)

    service = build('drive', 'v2', credentials=creds)
    return service

def retrieve_drive_files():
    service = auth()
    result = []
    page_token = None
    while True:
        try:
            param = {}
            if page_token:
                param['pageToken'] = page_token
            files = service.files().list(**param).execute()
            for file_ in files['items']:
                if '(respostas)' in file_['title'] and 'SEM' in file_['title']:
                    result.append(file_)

            page_token = files.get('nextPageToken')
            if not page_token:
                break
        except:
            print('An error occurred on @retrieve_drive_files')
            break
    for item in result:
        print(item['title'])
    return result

def download_csv_file(id, files=None):
    if files is None:
        files = retrieve_drive_files()
    service = auth()
    HEADERS = b'"xA","xB","xC","xD","Q01","Q02","Q03","Q04","Q05","Q06","Q07","Q08","Q09","Q10"'
    for item in files:
        if item['id'] == id:
            data = service.files().export(fileId=item['id'], mimeType='text/csv').execute()
            data = b'\n'.join([HEADERS] + data.split(b'\n')[1:])
            print('@download_csv_file', item['title'])
            return item['title'], io.BytesIO(data)

def load_csv(fn):
    df = pandas.read_csv(fn)
    df['xC'] = df['xC'].str.replace(' / 10', '').astype(float) 
    return df 

def top_students(df, ammount=10):
    df = df.sort_values(by=(['xC', 'xA']), ascending=[False, True])[:ammount]
    return df

def bottom_students(df, ammount=10):
    df = df.sort_values(by=(['xC', 'xA']), ascending=[True, True])[:ammount]
    return df

def calculate_statistics(df):
    statistics = {}
    statistics['MIN'] = np.min(df['xC'])
    statistics['MAX'] = np.max(df['xC'])
    statistics['MEAN'] = np.mean(df['xC'])
    statistics['STD'] = np.std(df['xC'])
    for question_id in ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10']:
        question_title = 'Q{}'.format(question_id)        
        answers = {'A': 0, 'B': 0, 'C': 0, 'D':0, 'E': 0}
        for _, student_data in df.iterrows():
            if student_data[question_title] in answers:
                answers[student_data[question_title]] += 1
            else:
                answers[student_data[question_title]] = 1
        statistics[question_title] = answers
    return statistics

def duplicate(ws, origin, destination):
    ws[destination].font = copy(ws[origin].font)
    ws[destination].border = copy(ws[origin].border)
    ws[destination].fill = copy(ws[origin].fill)
    ws[destination].number_format = copy(ws[origin].number_format)
    ws[destination].protection = copy(ws[origin].protection)
    ws[destination].alignment = copy(ws[origin].alignment)
    if type(ws[destination]).__name__ != 'MergedCell':
        ws[destination].value = copy(ws[origin].value)

    if (ws.row_dimensions[get_row(origin)].height == 7.5):
        ws.row_dimensions[get_row(destination)].height = 7.5

def generate_tbl(fn, fdata=None):
    print('@GEN TBL>>', fn, fdata)
    fn_formatted = fn.replace('inputs/', '').replace('(respostas)', '').replace('.csv', '').replace('.xlsx', '').replace('-', '').replace('  ', ' ').strip()
    df = load_csv(fdata if fdata else fn)
    print('GEN TBL>>SHAPE::', df.shape)
    print('GEN TBL>>GENERATING REPORT::{}'.format(fn_formatted))
    students = bottom_students(df, ammount=df.shape[0] - 20)
    stats = calculate_statistics(students)

    wb = openpyxl.load_workbook(filename='destrator/inputs/TEMPLATE_TBL.xlsx')
    ws = wb.active
    ws['D3'] = fn_formatted
    ws['G3'] = datetime.datetime.now().strftime('%d/%m/%Y')
    ws['C6'] = 'MAIOR: {:.2f}'.format(stats['MAX'])
    ws['D6'] = 'MENOR: {:.2f}'.format(stats['MIN'])
    ws['E6'] = 'MÉDIA: {:.2f}'.format(stats['MEAN'])
    ws['F6'] = 'DESVIO: {:.2f}'.format(stats['STD'])
        
    students = list(students.iterrows())
    if students:
        ws['G10'] = students[0][1]['xC']
        ws['C10'] = students[0][1]['xD'].upper()
        ws['B10'] = '1.  '
        ws['B10'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws['B10'].border = copy(ws['B5'].border)
        ws['C10'].border = copy(ws['C5'].border)
        ws['D10'].border = copy(ws['D5'].border)
        ws['E10'].border = copy(ws['E5'].border)
        ws['F10'].border = copy(ws['F5'].border)
        ws['G10'].border = copy(ws['G5'].border)
        for index, (_, student) in enumerate(students[1:]):
            row = str(11 + index)
            duplicate(ws, 'C10', 'C' + row)
            duplicate(ws, 'G10', 'G' + row)
            ws['C' + row].border = copy(ws['A1'].border)
            ws['B' + row].border = copy(ws['B5'].border)
            ws['G' + row].border = copy(ws['G5'].border)
            ws['G' + row] = student['xC']
            ws['C' + row] = student['xD'].upper()
            ws['B' + row] = '{}.  '.format(index + 2)
            ws['B' + row].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
            if type(ws['C' + row]).__name__ != 'MergedCell':
                ws.merge_cells('C{}:F{}'.format(row, row))
        ws['B' + str(9 + len(students))].border = copy(ws['B6'].border)
        ws['C' + str(9 + len(students))].border = copy(ws['C6'].border)
        ws['D' + str(9 + len(students))].border = copy(ws['D6'].border)
        ws['E' + str(9 + len(students))].border = copy(ws['E6'].border)
        ws['F' + str(9 + len(students))].border = copy(ws['F6'].border)
        ws['G' + str(9 + len(students))].border = copy(ws['G6'].border)

    wb.save('destrator/outputs/' + fn_formatted.upper() + '_TBL.xlsx')
    return 'destrator/outputs/' + fn_formatted.upper() + '_TBL.xlsx'

def generate_score_z(fn, fdata=None):
    print('@GEN SCORE Z>>', fn, fdata)
    fn_formatted = fn.replace('inputs/', '').replace('(respostas)', '').replace('.csv', '').replace('.xlsx', '').replace('-', '').replace('  ', ' ').strip()
    df = load_csv(fdata if fdata else fn)
    print('GEN SCORE Z>>SHAPE::', df.shape)
    print('GEN SCORE Z>>GENERATING REPORT::{}'.format(fn_formatted))
    students = top_students(df, ammount=20)
    stats = calculate_statistics(students)

    wb = openpyxl.load_workbook(filename='destrator/inputs/TEMPLATE_SCORE_Z.xlsx')
    ws = wb.active
    ws['D3'] = fn_formatted
    ws['G3'] = datetime.datetime.now().strftime('%d/%m/%Y')
    ws['C6'] = 'MAIOR: {:.2f}'.format(stats['MAX'])
    ws['D6'] = 'MENOR: {:.2f}'.format(stats['MIN'])
    ws['E6'] = 'MÉDIA: {:.2f}'.format(stats['MEAN'])
    ws['F6'] = 'DESVIO: {:.2f}'.format(stats['STD'])
        
    students = list(students.iterrows())
    if students:
        ws['G10'] = students[0][1]['xC']
        ws['C10'] = students[0][1]['xD'].upper()
        ws['B10'] = '1.  '
        ws['B10'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws['B10'].border = copy(ws['B5'].border)
        ws['C10'].border = copy(ws['C5'].border)
        ws['D10'].border = copy(ws['D5'].border)
        ws['E10'].border = copy(ws['E5'].border)
        ws['F10'].border = copy(ws['F5'].border)
        ws['G10'].border = copy(ws['G5'].border)
        for index, (_, student) in enumerate(students[1:]):
            row = str(11 + index)
            duplicate(ws, 'C10', 'C' + row)
            duplicate(ws, 'G10', 'G' + row)
            ws['C' + row].border = copy(ws['A1'].border)
            ws['B' + row].border = copy(ws['B5'].border)
            ws['G' + row].border = copy(ws['G5'].border)
            ws['G' + row] = student['xC']
            ws['C' + row] = student['xD'].upper()
            ws['B' + row] = '{}.  '.format(index + 2)
            ws['B' + row].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
            if type(ws['C' + row]).__name__ != 'MergedCell':
                ws.merge_cells('C{}:F{}'.format(row, row))
        ws['B' + str(9 + len(students))].border = copy(ws['B6'].border)
        ws['C' + str(9 + len(students))].border = copy(ws['C6'].border)
        ws['D' + str(9 + len(students))].border = copy(ws['D6'].border)
        ws['E' + str(9 + len(students))].border = copy(ws['E6'].border)
        ws['F' + str(9 + len(students))].border = copy(ws['F6'].border)
        ws['G' + str(9 + len(students))].border = copy(ws['G6'].border)

    wb.save('destrator/outputs/' + fn_formatted.upper() + '_SCORE_Z.xlsx')
    return 'destrator/outputs/' + fn_formatted.upper() + '_SCORE_Z.xlsx'

def generate_destrator(fn, fdata=None):
    print('@GEN DESTRATOR >>', fn, fdata)
    fn_formatted = fn.replace('inputs/', '').replace('(respostas)', '').replace('.csv', '').replace('.xlsx', '').replace('-', '').replace('  ', ' ').strip()
    df = load_csv(fdata if fdata else fn)
    print('GEN DESTRATOR>>SHAPE::', df.shape)
    print('GEN DESTRATOR>>GENERATING REPORT::{}'.format(fn_formatted))
    students = df
    xc = students['xC'].hist(bins=range(12), range=(0, 11), alpha=0.5, align='left', color='black')
    xc.set_ylabel('Frequência')
    xc.set_xlabel('Notas')
    xc.set_title('Frequência de Notas')
     # s is an instance of Series
    fig = xc.get_figure()
    fig.savefig('destrator/outputs/curve.png')
    pyplot.close(fig) 
    stats = calculate_statistics(students)

    wb = openpyxl.load_workbook(filename='destrator/inputs/TEMPLATE_DESTRATOR.xlsx')
    ws = wb.active
    ws['D3'] = fn_formatted
    ws['G3'] = datetime.datetime.now().strftime('%d/%m/%Y')
    ws['C6'] = 'MAIOR: {:.2f}'.format(stats['MAX'])
    ws['D6'] = 'MENOR: {:.2f}'.format(stats['MIN'])
    ws['E6'] = 'MÉDIA: {:.2f}'.format(stats['MEAN'])
    ws['F6'] = 'DESVIO: {:.2f}'.format(stats['STD'])
    
    # destractor matrix header
    ws.merge_cells('D{}:F{}'.format(9, 9))
    
    # destractor matrix
    for index, question_id in enumerate(['01', '02', '03', '04', '05', '06', '07', '08', '09', '10']):
        question_title = 'Q{}'.format(question_id)
        answers = ['A', 'B', 'C', 'D', 'E']
        for a_index, answer in enumerate(answers):
            ws[chr(ord('C') + a_index).upper() + str(11 + index)] = stats[question_title][answer]
    

    img = openpyxl.drawing.image.Image('destrator/outputs/curve.png')
    img.anchor = 'C22'
    ws.add_image(img)

    wb.save('destrator/outputs/' + fn_formatted.upper() + '_DESTRATOR.xlsx')
    return 'destrator/outputs/' + fn_formatted.upper() + '_DESTRATOR.xlsx'


if __name__ == '__main__':
    retrieve_drive_files()